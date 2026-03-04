#!/usr/bin/env python3
"""
Excel 讀取工具 — 支援 .xls / .xlsx
用法:
  python read-excel.py <file> [options]

選項:
  --sheet <name|index>    指定工作表（名稱或從 0 開始的索引）
  --sheets                列出所有工作表名稱
  --range <A1:Z10>        指定讀取範圍（如 A1:F20）
  --rows <start>-<end>    指定行範圍（如 1-50，從 1 開始）
  --cols <start>-<end>    指定列範圍（如 1-10，從 1 開始）
  --search <text>         搜尋包含特定文字的儲存格
  --header <row>          指定標題列（從 1 開始，預設 1）
  --max-rows <n>          最多輸出 n 行（預設 100）
  --max-cols <n>          最多輸出 n 列（預設 30）
  --tsv                   以 TSV 格式輸出（預設 markdown 表格）
  --raw                   不截斷長文字
"""
import sys
import os
import re
import argparse

def col_letter_to_num(letter):
    """A=1, B=2, ..., Z=26, AA=27, ..."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result

def parse_range(range_str):
    """解析 A1:F20 格式"""
    m = re.match(r'([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)', range_str)
    if not m:
        return None
    c1, r1, c2, r2 = m.groups()
    return (int(r1), col_letter_to_num(c1), int(r2), col_letter_to_num(c2))

def truncate(val, max_len=40):
    """截斷長文字"""
    s = str(val) if val is not None else ''
    s = s.replace('\n', ' ').replace('\r', '')
    if len(s) > max_len:
        return s[:max_len-2] + '..'
    return s

def read_xls(filepath):
    """讀取 .xls（xlrd）"""
    import xlrd
    wb = xlrd.open_workbook(filepath)
    return wb

def read_xlsx(filepath):
    """讀取 .xlsx（openpyxl）"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    return wb

def get_sheet_names(wb, is_xls):
    if is_xls:
        return wb.sheet_names()
    else:
        return wb.sheetnames

def get_sheet(wb, is_xls, sheet_id):
    names = get_sheet_names(wb, is_xls)
    if isinstance(sheet_id, int):
        if sheet_id < len(names):
            name = names[sheet_id]
        else:
            print(f"Error: sheet index {sheet_id} out of range (0-{len(names)-1})")
            sys.exit(1)
    else:
        name = sheet_id
        if name not in names:
            # Try partial match
            matches = [n for n in names if sheet_id.lower() in n.lower()]
            if matches:
                name = matches[0]
                print(f"[Matched sheet: {name}]")
            else:
                print(f"Error: sheet '{sheet_id}' not found. Available: {names}")
                sys.exit(1)
    if is_xls:
        return wb.sheet_by_name(name), name
    else:
        return wb[name], name

def get_cell_value(sheet, row, col, is_xls):
    """取得儲存格值（row/col 從 0 開始）"""
    try:
        if is_xls:
            return sheet.cell_value(row, col)
        else:
            cell = sheet.cell(row=row+1, column=col+1)
            return cell.value
    except (IndexError, ValueError):
        return None

def get_dimensions(sheet, is_xls):
    """取得工作表的行列數"""
    if is_xls:
        return sheet.nrows, sheet.ncols
    else:
        return sheet.max_row or 0, sheet.max_column or 0

def search_cells(sheet, is_xls, text, nrows, ncols):
    """搜尋包含特定文字的儲存格"""
    results = []
    text_lower = text.lower()
    for r in range(nrows):
        for c in range(ncols):
            val = get_cell_value(sheet, r, c, is_xls)
            if val is not None and text_lower in str(val).lower():
                results.append((r+1, c+1, str(val)[:80]))
    return results

def main():
    parser = argparse.ArgumentParser(description='Read Excel files (.xls/.xlsx)')
    parser.add_argument('file', help='Excel file path')
    parser.add_argument('--sheet', default=None, help='Sheet name or 0-based index')
    parser.add_argument('--sheets', action='store_true', help='List all sheet names')
    parser.add_argument('--range', default=None, help='Cell range (e.g. A1:F20)')
    parser.add_argument('--rows', default=None, help='Row range (e.g. 1-50)')
    parser.add_argument('--cols', default=None, help='Column range (e.g. 1-10)')
    parser.add_argument('--search', default=None, help='Search for text in cells')
    parser.add_argument('--header', type=int, default=1, help='Header row number (1-based)')
    parser.add_argument('--max-rows', type=int, default=100, help='Max rows to output')
    parser.add_argument('--max-cols', type=int, default=30, help='Max columns to output')
    parser.add_argument('--tsv', action='store_true', help='Output as TSV')
    parser.add_argument('--raw', action='store_true', help='Do not truncate text')
    args = parser.parse_args()

    filepath = args.file
    if not os.path.exists(filepath):
        print(f"Error: file not found: {filepath}")
        sys.exit(1)

    is_xls = filepath.lower().endswith('.xls')
    is_xlsx = filepath.lower().endswith('.xlsx') or filepath.lower().endswith('.xlsm')

    if not (is_xls or is_xlsx):
        print(f"Error: unsupported format. Use .xls or .xlsx")
        sys.exit(1)

    try:
        if is_xls:
            wb = read_xls(filepath)
        else:
            wb = read_xlsx(filepath)
    except Exception as e:
        print(f"Error opening file: {e}")
        sys.exit(1)

    # List sheets
    names = get_sheet_names(wb, is_xls)
    if args.sheets:
        print(f"Sheets ({len(names)}):")
        for i, name in enumerate(names):
            print(f"  [{i}] {name}")
        return

    # Select sheet
    sheet_id = args.sheet
    if sheet_id is not None:
        try:
            sheet_id = int(sheet_id)
        except ValueError:
            pass
    else:
        sheet_id = 0

    sheet, sheet_name = get_sheet(wb, is_xls, sheet_id)
    nrows, ncols = get_dimensions(sheet, is_xls)
    print(f"[Sheet: {sheet_name} | {nrows} rows x {ncols} cols]")

    # Search mode
    if args.search:
        results = search_cells(sheet, is_xls, args.search, nrows, ncols)
        print(f"\nFound {len(results)} matches for '{args.search}':")
        for r, c, val in results[:50]:
            print(f"  Row {r}, Col {c}: {val}")
        return

    # Determine range
    r_start, r_end = 0, nrows - 1
    c_start, c_end = 0, ncols - 1

    if args.range:
        parsed = parse_range(args.range)
        if parsed:
            r_start, c_start, r_end, c_end = parsed[0]-1, parsed[1]-1, parsed[2]-1, parsed[3]-1

    if args.rows:
        parts = args.rows.split('-')
        r_start = int(parts[0]) - 1
        r_end = int(parts[1]) - 1 if len(parts) > 1 else r_start

    if args.cols:
        parts = args.cols.split('-')
        c_start = int(parts[0]) - 1
        c_end = int(parts[1]) - 1 if len(parts) > 1 else c_start

    # Clamp
    r_end = min(r_end, r_start + args.max_rows - 1, nrows - 1)
    c_end = min(c_end, c_start + args.max_cols - 1, ncols - 1)

    max_text = 200 if args.raw else 40

    # Read data
    rows_data = []
    for r in range(r_start, r_end + 1):
        row = []
        for c in range(c_start, c_end + 1):
            val = get_cell_value(sheet, r, c, is_xls)
            row.append(truncate(val, max_text))
        rows_data.append(row)

    if not rows_data:
        print("(no data)")
        return

    # Output
    if args.tsv:
        for i, row in enumerate(rows_data):
            print(f"R{r_start+i+1}\t" + '\t'.join(row))
    else:
        # Markdown table
        num_cols = len(rows_data[0]) if rows_data else 0
        # Calculate column widths
        widths = [3] * num_cols
        for row in rows_data:
            for j, val in enumerate(row):
                widths[j] = max(widths[j], min(len(val), 40))

        for i, row in enumerate(rows_data):
            prefix = f"R{r_start+i+1:>4}"
            cells = [val.ljust(widths[j])[:widths[j]] for j, val in enumerate(row)]
            print(f"{prefix} | {'|'.join(cells)} |")
            if i == args.header - r_start - 1:
                sep = ['-' * widths[j] for j in range(num_cols)]
                print(f"{'':>5} | {'|'.join(sep)} |")

    remaining = nrows - 1 - r_end
    if remaining > 0:
        print(f"\n[... {remaining} more rows not shown. Use --rows or --max-rows to see more]")

    if not is_xls:
        wb.close()

if __name__ == '__main__':
    main()
